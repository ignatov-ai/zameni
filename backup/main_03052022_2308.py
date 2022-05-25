import sys
from datetime import datetime

from PyQt6 import QtGui
from PyQt6.QtCore import QDate
from PyQt6.QtWidgets import (QApplication, QLabel, QMainWindow,
    QPushButton, QTabWidget, QWidget, QLineEdit, QComboBox, QDateEdit)

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from openpyxl.workbook import Workbook

headers = ['Фамилия', 'Имя', 'Отчество', 'Таб. номер', 'Должность', 'Дата открытия больничного листа', 'Дата закрытия больничного листа']
workbook_name = 'test.xlsx'

'''
wb = Workbook()
page = wb.active
page.title = 'Учет больничных листов'
page.append(headers) # write the headers to the first line
'''

eng_chars = u"~!@#$%^&qwertyuiop[]asdfghjkl;'zxcvbnm,./QWERTYUIOP{}ASDFGHJKL:\"|ZXCVBNM<>?"
rus_chars = u"ё!\"№;%:?йцукенгшщзхъфывапролджэячсмитьбю.ЙЦУКЕНГШЩЗХЪФЫВАПРОЛДЖЭ/ЯЧСМИТЬБЮ,"
trans_table = dict(zip(eng_chars, rus_chars))

# функция исправления неправильной раскладки
def fix_input(s):
    return u''.join([trans_table.get(c, c) for c in s])

# определение текущей даты
day = datetime.now().day
if day < 10:
    day = str('0')+str(day)
month = datetime.now().month
if month < 10:
    month = str('0')+str(month)
today = day +'.'+ month +'.'+ str(datetime.now().year)
print(today)
current_day = datetime.today().weekday()

# выгрузка БД с расписанием
raspisanie = []
with open('../raspisanie_done.csv', 'r') as url:
  for line in url:
    raspisanie.append(line.strip().split(';'))

'''
for j in range(len(raspisanie)):
    print(raspisanie[j][0] + '. ' + raspisanie[j][1] +' '+ raspisanie[j][2] + ' ', end='')
    for i in range(3+current_day*10, 3+(current_day+1)*10):
        print(str(i-2-current_day*10) + '. '+raspisanie[j][i]+' | ', end='')
    print()
'''

# выгрузка БД с сотрудниками
sotrudniki = []
with open('../sotrudniki.csv', 'r') as url:
  for line in url:
    sotrudniki.append(line.strip().split(';'))

sotrudniki_fio = []
for i in range(len(sotrudniki)):
    s = sotrudniki[i][3]+'. '+sotrudniki[i][0]+' '+sotrudniki[i][1]+' '+sotrudniki[i][2]
    sotrudniki_fio.append(s)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Вкладки для замен")

        self.tabs = QTabWidget()
        tab1 = QWidget()
        tab2 = QWidget()
        tab3 = QWidget()
        tab4 = QWidget()
        self.tabs.setMovable(True)
        self.tabs.addTab(tab1, 'Создание больничного листа')
        self.tabs.addTab(tab2, 'Создание замены')
        self.tabs.addTab(tab3, 'Журнал больничных листов')
        self.tabs.addTab(tab4, 'Журнал замен')
        self.setCentralWidget(self.tabs)

        ###################################################################################
        ######################### ВКЛАДКА СОЗДАНИЯ БОЛЬНИЧНОГО ЛИСТА ######################
        ###################################################################################

        # поле поиска и выбора учителя
        lbl_find = QLabel(tab1)
        lbl_find.setText('Поиск: ')
        lbl_find.move(30,60)
        lbl_find.setFixedWidth(70)

        teach_find = QLineEdit(tab1)
        teach_find.setFocus()
        teach_find.textChanged.connect(self.teachFind)
        teach_find.move(100,60)
        teach_find.setFixedWidth(200)

        self.teach_select = QComboBox(tab1)
        self.teach_select.move(30,100)
        self.teach_select.addItems(sotrudniki_fio)
        self.teach_select.setFixedWidth(270)
        self.teach_select.currentIndex()

        # выбор даты начала замены
        self.zamena_dateStart = QDateEdit(tab1, calendarPopup=True)
        self.zamena_dateStart.move(100,140)
        self.zamena_dateStart.setFixedWidth(200)
        self.zamena_dateStart.setDate(QDate.currentDate())

        lbl_zamenaStart = QLabel(tab1)
        lbl_zamenaStart.setText('Начало: ')
        lbl_zamenaStart.move(30,142)
        lbl_zamenaStart.setFixedWidth(70)

        # выбор даты окончания замены
        self.zamena_dateEnd = QDateEdit(tab1, calendarPopup=True)
        self.zamena_dateEnd.move(100,180)
        self.zamena_dateEnd.setFixedWidth(200)
        self.zamena_dateEnd.setDate(QDate.currentDate())

        lbl_zamenaEnd = QLabel(tab1)
        lbl_zamenaEnd.move(30,182)
        lbl_zamenaEnd.setText('Окончание: ')
        lbl_zamenaEnd.setFixedWidth(70)

        # кнопки отмены и добавления замены в журнал
        okButton = QPushButton(tab1)
        okButton.setText("Добавить")
        okButton.move(30,220)
        okButton.setFixedWidth(100)
        okButton.clicked.connect(self.zamena_add)
        #okButton.addAction(self.zamena_add)
        cancelButton = QPushButton(tab1)
        cancelButton.setText("Назад")
        cancelButton.move(200,220)
        cancelButton.setFixedWidth(100)

        ###################################################################################
        ############################## ВКЛАДКА СОЗДАНИЯ ЗАМЕНЫ ############################
        ###################################################################################

        # выбор учителя для замены
        lbl_find_2 = QLabel(tab2)
        lbl_find_2.setText('Поиск: ')
        lbl_find_2.move(30, 60)
        lbl_find_2.setFixedWidth(70)

        teach_find_2 = QLineEdit(tab2)
        teach_find_2.setFocus()
        teach_find_2.textChanged.connect(self.teachFind_2)
        teach_find_2.move(150, 60)
        teach_find_2.setFixedWidth(200)

        self.teach_select_2 = QComboBox(tab2)
        self.teach_select_2.move(30, 100)
        self.teach_select_2.addItems(sotrudniki_fio)
        self.teach_select_2.setFixedWidth(270)
        self.teach_select_2.currentIndex()
        self.teach_select_2.textActivated.connect(self.zamena_teach_select)

        self.lbl_teach_select = QLabel(tab2)
        self.lbl_teach_select.setText('Выбранный учитель: ')
        self.lbl_teach_select.move(30, 142)
        self.lbl_teach_select.setFixedWidth(120)

        self.lbl_teach_select_2 = QLabel(tab2)
        self.lbl_teach_select_2.setText('Учитель еще не выбран!')
        self.lbl_teach_select_2.move(150, 142)
        self.lbl_teach_select_2.setFixedWidth(250)

        # выбор даты для замены
        lbl_zamena_2 = QLabel(tab2)
        lbl_zamena_2.setText('Дата замены: ')
        lbl_zamena_2.move(30, 182)
        lbl_zamena_2.setFixedWidth(70)

        self.zamena_select_2 = QDateEdit(tab2, calendarPopup=True)
        self.zamena_select_2.move(150, 180)
        self.zamena_select_2.setFixedWidth(200)
        self.zamena_select_2.setDate(QDate.currentDate())
        self.zamena_select_2.dateChanged.connect(self.zamena_date_select)

        self.lbl_zamena_s = QLabel(tab2)
        self.lbl_zamena_s.setText('Выбранная дата: ')
        self.lbl_zamena_s.move(30, 222)
        self.lbl_zamena_s.setFixedWidth(100)

        self.lbl_zamena_s_2 = QLabel(tab2)
        #self.lbl_zamena_s_2.setText(today)
        self.lbl_zamena_s_2.setText('Дата еще не выбрана!')
        self.lbl_zamena_s_2.move(150, 222)
        self.lbl_zamena_s_2.setFixedWidth(150)

        zamenaButton = QPushButton(tab2)
        zamenaButton.setText("Построить замены для выбранной даты")
        zamenaButton.move(30, 260)
        zamenaButton.setFixedWidth(320)
        zamenaButton.clicked.connect(self.zamena_add_form)

    # вывод в выпадающий список учителей по фильтру из поля поиска
    def teachFind(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find.append(sotrudniki_fio[i])
        if len(sotrudniki_find) > 0:
            self.teach_select.clear()
            self.teach_select.addItems(sotrudniki_find)
        else:
            self.teach_select.addItems(sotrudniki_fio)

    def teachFind_2(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find_2 = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6+lenFioSrez] == fioSrez:
                sotrudniki_find_2.append(sotrudniki_fio[i])
        if len(sotrudniki_find_2) > 0:
            self.teach_select_2.clear()
            self.teach_select_2.addItems(sotrudniki_find_2)
        else:
            self.teach_select_2.addItems(sotrudniki_fio)

    def zamena_add(self):
        fio_text = self.teach_select.currentText()
        id = fio_text[:4]
        for i in range (len(sotrudniki)):
            if sotrudniki[i][3] == id:
                break
        zamena_zapis = [sotrudniki[i][j] for j in range(5)]
        zamena_zapis.append(self.zamena_dateStart.text())
        zamena_zapis.append(self.zamena_dateEnd.text())
        print(zamena_zapis)

        wb = load_workbook(filename=workbook_name)
        ws = wb['Учет больничных листов']
        ws.append(zamena_zapis)
        sheet = wb.active
        # выравнивание столбцов D E F G по центру
        for c in 'DEFG':
            currRow = c+str(sheet.max_row)
            cell = sheet[str(currRow)]
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        wb.save(filename=workbook_name)
        wb.close()

    def zamena_date_select(self):
        selected_date = self.zamena_select_2.text()
        #print(selected_date)
        self.lbl_zamena_s_2.setText(selected_date)
        return selected_date

    def zamena_teach_select(self):
        selected_fio_text = self.teach_select_2.currentText()
        #print(selected_fio_text)
        self.lbl_teach_select_2.setText(selected_fio_text)
        return selected_fio_text[6:]

    def zamena_add_form(self):
        day, month, year = (int(x) for x in self.zamena_select_2.text().split('.'))
        sel_date = datetime.weekday(datetime(year, month, day))
        sel_teach = self.teach_select_2.currentText()[:4]

        for teach_id in range(len(sotrudniki)):
            if raspisanie[teach_id][0] == sel_teach:
                break
        # вывод заменяемого учителя и расписание его уроков на этот день
        print(raspisanie[teach_id][0] + '. ' + raspisanie[teach_id][1] + ' ' + raspisanie[teach_id][2] + ' ', end='')
        for i in range(4 + sel_date * 10, 4 + (sel_date + 1) * 10):
            print(str(i - 3 - sel_date * 10) + '. ' + raspisanie[teach_id][i] + ' | ', end='')
        print()
        sel_predmet = raspisanie[teach_id][2]
        # вывод учителей ТОГО ЖЕ предмета
        kandidati = []
        for i in range(len(raspisanie)):
            if raspisanie[i][2] == sel_predmet and raspisanie[i][0] != sel_teach:
                kandidati.append(raspisanie[i][0]+'. '+raspisanie[i][1])

        # вывод учителей ОСТАЛЬНЫХ предметов
        for i in range(len(raspisanie)):
            if raspisanie[i][2] != sel_predmet:
                kandidati.append(raspisanie[i][0]+'. '+raspisanie[i][1])

        n_lessons = 10 - raspisanie[teach_id][4 + sel_date * 10: 4 + (sel_date + 1) * 10].count('-')
        #print(n_lessons,kandidati)




app = QApplication(sys.argv)
app.setWindowIcon(QtGui.QIcon('../icon.png'))

window = MainWindow()
window.setGeometry(400, 200, 800, 600)
window.setWindowTitle('Составитель замен. Текущая дата: ' + today)
window.show()

app.exec()