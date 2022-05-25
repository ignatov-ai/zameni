import sys
from datetime import datetime

from PyQt6 import QtGui
from PyQt6.QtCore import QDate
from PyQt6.QtWidgets import (QApplication, QLabel, QMainWindow,
                             QPushButton, QTabWidget, QWidget, QLineEdit, QComboBox, QDateEdit)

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# from openpyxl.workbook import Workbook

headers = ['Фамилия', 'Имя', 'Отчество', 'Таб. номер', 'Должность', 'Дата открытия больничного листа',
           'Дата закрытия больничного листа']
bolnichniy_book_name = 'test.xlsx'
zameni_book_name = 'zameni.xlsx'

'''
wb = Workbook()
page = wb.active
page.title = 'Учет больничных листов'
page.append(headers) # write the headers to the first line
'''

eng_chars = u"~!@#$%^&qwertyuiop[]asdfghjkl;'zxcvbnm,./QWERTYUIOP{}ASDFGHJKL:\"|ZXCVBNM<>?"
rus_chars = u"ё!\"№;%:?йцукенгшщзхъфывапролджэячсмитьбю.ЙЦУКЕНГШЩЗХЪФЫВАПРОЛДЖЭ/ЯЧСМИТЬБЮ,"
trans_table = dict(zip(eng_chars, rus_chars))


# функция исправления неправильной раскладки
def fix_input(st):
    return u''.join([trans_table.get(c, c) for c in st])


# определение текущей даты
day = datetime.now().day
if day < 10:
    day = str('0') + str(day)
month = datetime.now().month
if month < 10:
    month = str('0') + str(month)
today = day + '.' + month + '.' + str(datetime.now().year)
print(today)
current_day = datetime.today().weekday()

# выгрузка БД с расписанием
raspisanie = []
with open('../raspisanie_done.csv', 'r') as url:
    for line in url:
        raspisanie.append(line.strip().split(';'))

'''
for j in range(len(raspisanie)):
    print(raspisanie[j][0] + '. ' + raspisanie[j][1] +' '+ raspisanie[j][2] + ' ', end='')
    for i in range(3+current_day*10, 3+(current_day+1)*10):
        print(str(i-2-current_day*10) + '. '+raspisanie[j][i]+' | ', end='')
    print()
'''

# выгрузка БД с сотрудниками
sotrudniki = []
with open('../sotrudniki.csv', 'r') as url:
    for line in url:
        sotrudniki.append(line.strip().split(';'))

sotrudniki_fio = []
for i in range(len(sotrudniki)):
    s = sotrudniki[i][3] + '. ' + sotrudniki[i][0] + ' ' + sotrudniki[i][1] + ' ' + sotrudniki[i][2]
    sotrudniki_fio.append(s)

class zamenaAddBtn(QPushButton):
    def __init__(self):
        QPushButton.__init__(self)
        self.setStyleSheet("background-color: rgb(255,255,255);")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Вкладки для замен")

        self.tabs = QTabWidget()
        self.tab1 = QWidget()
        tab2 = QWidget()
        self.tab3 = QWidget()
        self.tab4 = QWidget()
        self.tabs.setMovable(True)
        self.tabs.addTab(self.tab1, 'Создание больничного листа')
        self.tabs.addTab(tab2, 'Создание замены')
        self.tabs.addTab(self.tab3, 'Журнал больничных листов')
        self.tabs.addTab(self.tab4, 'Журнал замен')
        self.setCentralWidget(self.tabs)

        ###################################################################################
        ######################### ВКЛАДКА СОЗДАНИЯ БОЛЬНИЧНОГО ЛИСТА ######################
        ###################################################################################

        # поле поиска и выбора учителя
        lbl_find = QLabel(self.tab1)
        lbl_find.setText('Поиск: ')
        lbl_find.move(30, 60)
        lbl_find.setFixedWidth(70)

        teach_find = QLineEdit(self.tab1)
        teach_find.setFocus()
        teach_find.textChanged.connect(self.teachFind)
        teach_find.move(100, 60)
        teach_find.setFixedWidth(200)

        self.teach_select = QComboBox(self.tab1)
        self.teach_select.move(30, 100)
        self.teach_select.addItems(sotrudniki_fio)
        self.teach_select.setFixedWidth(270)
        self.teach_select.currentIndex()

        # выбор даты начала замены
        self.zamena_dateStart = QDateEdit(self.tab1, calendarPopup=True)
        self.zamena_dateStart.move(100, 140)
        self.zamena_dateStart.setFixedWidth(200)
        self.zamena_dateStart.setDate(QDate.currentDate())

        lbl_zamenaStart = QLabel(self.tab1)
        lbl_zamenaStart.setText('Начало: ')
        lbl_zamenaStart.move(30, 142)
        lbl_zamenaStart.setFixedWidth(70)

        # выбор даты окончания замены
        self.zamena_dateEnd = QDateEdit(self.tab1, calendarPopup=True)
        self.zamena_dateEnd.move(100, 180)
        self.zamena_dateEnd.setFixedWidth(200)
        self.zamena_dateEnd.setDate(QDate.currentDate())

        lbl_zamenaEnd = QLabel(self.tab1)
        lbl_zamenaEnd.move(30, 182)
        lbl_zamenaEnd.setText('Окончание: ')
        lbl_zamenaEnd.setFixedWidth(70)

        # кнопки отмены и добавления замены в журнал
        okButton = QPushButton(self.tab1)
        okButton.setText("Добавить")
        okButton.move(30, 220)
        okButton.setFixedWidth(100)
        okButton.clicked.connect(self.bolnichniy_add)
        cancelButton = QPushButton(self.tab1)
        cancelButton.setText("Назад")
        cancelButton.move(200, 220)
        cancelButton.setFixedWidth(100)

        ###################################################################################
        ############################## ВКЛАДКА СОЗДАНИЯ ЗАМЕНЫ ############################
        ###################################################################################

        # выбор учителя для замены
        lbl_find_2 = QLabel(tab2)
        lbl_find_2.setText('Поиск: ')
        lbl_find_2.move(30, 60)
        lbl_find_2.setFixedWidth(70)

        teach_find_2 = QLineEdit(tab2)
        teach_find_2.setFocus()
        teach_find_2.textChanged.connect(self.teachFind_2)
        teach_find_2.move(150, 60)
        teach_find_2.setFixedWidth(200)

        self.teach_select_2 = QComboBox(tab2)
        self.teach_select_2.move(30, 100)
        self.teach_select_2.addItems(sotrudniki_fio)
        self.teach_select_2.setFixedWidth(270)
        self.teach_select_2.currentIndex()
        self.teach_select_2.textActivated.connect(self.zamena_teach_select)

        self.lbl_teach_select = QLabel(tab2)
        self.lbl_teach_select.setText('Выбранный учитель: ')
        self.lbl_teach_select.move(30, 142)
        self.lbl_teach_select.setFixedWidth(120)

        self.lbl_teach_select_2 = QLabel(tab2)
        self.lbl_teach_select_2.setText('Учитель еще не выбран!')
        self.lbl_teach_select_2.move(150, 142)
        self.lbl_teach_select_2.setFixedWidth(250)

        # выбор даты для замены
        lbl_zamena_2 = QLabel(tab2)
        lbl_zamena_2.setText('Дата замены: ')
        lbl_zamena_2.move(30, 182)
        lbl_zamena_2.setFixedWidth(70)

        self.zamena_select_2 = QDateEdit(tab2, calendarPopup=True)
        self.zamena_select_2.move(150, 180)
        self.zamena_select_2.setFixedWidth(200)
        self.zamena_select_2.setDate(QDate.currentDate())
        self.zamena_select_2.dateChanged.connect(self.zamena_date_select)

        self.lbl_zamena_s = QLabel(tab2)
        self.lbl_zamena_s.setText('Выбранная дата: ')
        self.lbl_zamena_s.move(30, 222)
        self.lbl_zamena_s.setFixedWidth(100)

        self.lbl_zamena_s_2 = QLabel(tab2)
        # self.lbl_zamena_s_2.setText(today)
        self.lbl_zamena_s_2.setText('Дата еще не выбрана!')
        self.lbl_zamena_s_2.move(150, 222)
        self.lbl_zamena_s_2.setFixedWidth(150)

        zamenaButton = QPushButton(tab2)
        zamenaButton.setText("Построить замены для выбранной даты")
        zamenaButton.move(30, 260)
        zamenaButton.setFixedWidth(320)
        zamenaButton.clicked.connect(lambda ch, tab = tab2: self.zamena_lessons_build(tab))

        '''
        # построение списка уроков заменяемого учителя
        self.les_1 = '----------'
        self.les_2 = '----------'
        self.les_3 = '----------'
        self.les_4 = '----------'
        self.les_5 = '----------'
        self.les_6 = '----------'
        self.les_7 = '----------'
        self.les_8 = '----------'
        self.les_9 = '----------'
        self.les_10 = '----------'

        self.lbl_les_1_label = QLabel(tab2)
        self.lbl_les_1_label.setText('1.     ' + self.les_1)
        self.lbl_les_1_label.move(440, 62)
        self.lbl_les_1_label.setFixedWidth(200)

        self.lbl_les_2_label = QLabel(tab2)
        self.lbl_les_2_label.setText('2.     ' + self.les_2)
        self.lbl_les_2_label.move(440, 92)
        self.lbl_les_2_label.setFixedWidth(200)

        self.lbl_les_3_label = QLabel(tab2)
        self.lbl_les_3_label.setText('3.     ' + self.les_3)
        self.lbl_les_3_label.move(440, 122)
        self.lbl_les_3_label.setFixedWidth(200)

        self.lbl_les_4_label = QLabel(tab2)
        self.lbl_les_4_label.setText('4.     ' + self.les_4)
        self.lbl_les_4_label.move(440, 152)
        self.lbl_les_4_label.setFixedWidth(200)

        self.lbl_les_5_label = QLabel(tab2)
        self.lbl_les_5_label.setText('5.     ' + self.les_5)
        self.lbl_les_5_label.move(440, 182)
        self.lbl_les_5_label.setFixedWidth(200)

        self.lbl_les_6_label = QLabel(tab2)
        self.lbl_les_6_label.setText('6.     ' + self.les_6)
        self.lbl_les_6_label.move(440, 212)
        self.lbl_les_6_label.setFixedWidth(200)

        self.lbl_les_7_label = QLabel(tab2)
        self.lbl_les_7_label.setText('7.     ' + self.les_7)
        self.lbl_les_7_label.move(440, 242)
        self.lbl_les_7_label.setFixedWidth(200)

        self.lbl_les_8_label = QLabel(tab2)
        self.lbl_les_8_label.setText('8.     ' + self.les_8)
        self.lbl_les_8_label.move(440, 272)
        self.lbl_les_8_label.setFixedWidth(200)

        self.lbl_les_9_label = QLabel(tab2)
        self.lbl_les_9_label.setText('9.     ' + self.les_9)
        self.lbl_les_9_label.move(440, 302)
        self.lbl_les_9_label.setFixedWidth(200)

        self.lbl_les_10_label = QLabel(tab2)
        self.lbl_les_10_label.setText('10.   ' + self.les_10)
        self.lbl_les_10_label.move(440, 332)
        self.lbl_les_10_label.setFixedWidth(200)

        for i in range (10):
            self.les_zamena_add = QPushButton(tab2)
            self.les_zamena_add.setText("Добавить замену")
            self.les_zamena_add.move(650, 60 + i*30)
            self.les_zamena_add.setFixedWidth(120)
            self.les_zamena_add.setEnabled(False)
            self.les_zamena_add.clicked.connect(lambda ch, num = i + 1: self.zamena_add_form(num))
        '''

    # вывод в выпадающий список учителей по фильтру из поля поиска
    def teachFind(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find.append(sotrudniki_fio[i])
        if len(sotrudniki_find) > 0:
            self.teach_select.clear()
            self.teach_select.addItems(sotrudniki_find)
        else:
            self.teach_select.addItems(sotrudniki_fio)

    def teachFind_2(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find_2 = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find_2.append(sotrudniki_fio[i])
        if len(sotrudniki_find_2) > 0:
            self.teach_select_2.clear()
            self.teach_select_2.addItems(sotrudniki_find_2)
        else:
            self.teach_select_2.addItems(sotrudniki_fio)

    def bolnichniy_add(self):
        fio_text = self.teach_select.currentText()
        id = fio_text[:4]
        for i in range(len(sotrudniki)):
            if sotrudniki[i][3] == id:
                break
        zamena_zapis = [sotrudniki[i][j] for j in range(5)]
        zamena_zapis.append(self.zamena_dateStart.text())
        zamena_zapis.append(self.zamena_dateEnd.text())
        print(zamena_zapis)

        wb = load_workbook(filename=bolnichniy_book_name)
        ws = wb['Учет больничных листов']
        ws.append(zamena_zapis)
        sheet = wb.active
        # выравнивание столбцов D E F G по центру
        for c in 'DEFG':
            currRow = c + str(sheet.max_row)
            cell = sheet[str(currRow)]
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        wb.save(filename=bolnichniy_book_name)
        wb.close()

    def zamena_date_select(self):
        selected_date = self.zamena_select_2.text()
        # print(selected_date)
        self.lbl_zamena_s_2.setText(selected_date)
        return selected_date

    def zamena_teach_select(self):
        selected_fio_text = self.teach_select_2.currentText()
        # print(selected_fio_text)
        self.lbl_teach_select_2.setText(selected_fio_text)
        return selected_fio_text[6:]

    def zamena_lessons_build(self,tab2):
        day, month, year = (int(x) for x in self.zamena_select_2.text().split('.'))
        date = datetime(year, month, day)
        sel_date = datetime.weekday(datetime(year, month, day))
        sel_teach_all = self.teach_select_2.currentText()
        sel_teach = self.teach_select_2.currentText()[:4]

        for teach_id in range(len(sotrudniki)):
            if raspisanie[teach_id][0] == sel_teach:
                break

        for sotr_teach_id in range(len(sotrudniki)):
            if sotrudniki[sotr_teach_id][3] == sel_teach:
                break
        print('Индекс выбранного учителя в массиве сотрудников:', teach_id)

        self.les_zamena_add_buttons = [zamenaAddBtn() for i in range(10)]

        for i in range(4 + sel_date * 10, 4 + (sel_date + 1) * 10):
            print(str(i - 3 - sel_date * 10) + '. ' + raspisanie[teach_id][i] + ' | ', end='')

            self.lbl_les_1_label = QLabel(tab2)
            self.lbl_les_1_label.setText('1.     ' + raspisanie[teach_id][i])
            self.lbl_les_1_label.move(440, 62 + i * 30)
            self.lbl_les_1_label.setFixedWidth(200)


            les_zamena_add.setText("Добавить замену")
            les_zamena_add.move(650, 60 + i * 30)
            les_zamena_add.setFixedWidth(120)
            les_zamena_add.clicked.connect(lambda ch, num=i + 1: self.zamena_add_form(num))

            if raspisanie[teach_id][i] == '-':
                les_zamena_add.setEnabled(False)
                print('Конпка НЕ работает')
            else:
                les_zamena_add.setEnabled(True)
                print('Конпка работает')

    def zamena_add_form(self, i):
        num_les = i
        day, month, year = (int(x) for x in self.zamena_select_2.text().split('.'))
        date = datetime(year, month, day)
        sel_date = datetime.weekday(datetime(year, month, day))
        sel_teach_all = self.teach_select_2.currentText()
        sel_teach = self.teach_select_2.currentText()[:4]

        if sel_date == 5 or sel_date == 6:
            self.lbl_zamena_s_2.setText('<p style="color: rgb(250, 55, 55);">ВЫБРАНА НЕВЕРНАЯ ДАТА</p>', )
        else:
            for teach_id in range(len(sotrudniki)):
                if raspisanie[teach_id][0] == sel_teach:
                    break

            for sotr_teach_id in range(len(sotrudniki)):
                if sotrudniki[sotr_teach_id][3] == sel_teach:
                    break
            print('Индекс выбранного учителя в массиве сотрудников:', teach_id)

            # вывод заменяемого учителя и расписание его уроков на этот день
            sel_teach_num_les = []
            #print(raspisanie[teach_id][0] + '. ' + raspisanie[teach_id][1] + ' '
            #+ raspisanie[teach_id][2] + ' ', end='')
            for i in range(4 + sel_date * 10, 4 + (sel_date + 1) * 10):
                # print(str(i - 3 - sel_date * 10) + '. ' + raspisanie[teach_id][i] + ' | ', end='')
                if raspisanie[teach_id][i] != '-':
                    # сохранение в массив sel_teach_num_les уроков для замены (окна пропущены)
                    sel_teach_num_les.append(i)
            print('Номера уроков для замены:', *sel_teach_num_les)
            # print()
            self.sel_predmet = raspisanie[teach_id][2]
            # вывод учителей ТОГО ЖЕ предмета
            kandidati = []
            self.send_teach_id = teach_id

            for j in range(len(raspisanie)):
                for i in sel_teach_num_les:
                    if raspisanie[j][2] == self.sel_predmet and raspisanie[j][0] != sel_teach \
                            and raspisanie[j][i] == '-':
                        kand_temp = str(i - 3 - sel_date * 10) + ';' + str(raspisanie[j][0]) + ';' \
                                    + str(raspisanie[j][1]) + ';' + str(raspisanie[teach_id][i])
                        kandidati.append(kand_temp)
                        # print(raspisanie[j][0]+'. '+raspisanie[j][1], end=' | ')
                # print()

            # вывод учителей ОСТАЛЬНЫХ предметов
            for j in range(len(raspisanie)):
                for i in sel_teach_num_les:
                    if raspisanie[j][2] != self.sel_predmet and raspisanie[j][i] == '-':
                        kand_temp = str(i - 3 - sel_date * 10) + ';' + str(raspisanie[j][0]) + ';' + \
                                    str(raspisanie[j][1]) + ';' + str(raspisanie[teach_id][i])
                        kandidati.append(kand_temp)
                        # print(raspisanie[j][0]+'. '+raspisanie[j][1], end=' | ')
                # print()
            # print(kandidati)

            kandidati = sorted(kandidati, key=lambda row: row[0])
            n_lessons = 10 - raspisanie[teach_id][4 + sel_date * 10: 4 + (sel_date + 1) * 10].count('-')
            print('Количество заменяемых уроков:', n_lessons)
            print('Количество кандидатов для замены ВСЕХ уроков:', len(kandidati))

            #num_les = 6
            self.win_zamena = zamena_add_window(teach_id, num_les, sel_teach_all, kandidati,
                                                date, sel_date, self.sel_predmet)
            self.win_zamena.show()


class zamena_add_window(QWidget):
    def __init__(self, teach_id, num_les, sel_teach, kandidati, date, sel_weekday, sel_predmet):
        super().__init__()
        self.selected_teacher_fio = sel_teach[6:]
        self.selected_teacher_num_tab = sel_teach[:4]
        self.setWindowTitle('Окно добавления замены')
        self.setFixedSize(360, 340)
        year, month, day = str(date)[:10].split('-')
        self.sel_date = day + '.' + month + '.' + year
        self.zamena_klass = raspisanie[teach_id][4 + int(sel_weekday) * 10 + int(num_les)]
        self.predmet = sel_predmet
        print('Выбранная дата:', self.sel_date)
        print('Выбранный учитель ТАБ №:', self.selected_teacher_num_tab)
        print('Выбранный учитель ФИО:', self.selected_teacher_fio)
        print('Кафедра на котором работает учитель:', self.predmet)
        print('Номер дня недели:', sel_weekday + 1)
        print('Номер заменяемого урока:', num_les)
        print('Индекс столбца заменяемого урока:', 4 + int(sel_weekday) * 10 + int(num_les))
        print('Класс для которого выставляется замена на данном уроке:', self.zamena_klass)

        kandidati_list = []
        for i in range(len(kandidati)):
            sel_les, num_tab_kand, fio_kand, zamena_les = list(map(str, kandidati[i].split(';')))
            kandidati_list.append(num_tab_kand + '. ' + fio_kand)

        # поле поиска и выбора учителя
        self.lbl_find = QLabel(self)
        self.lbl_find.setText('Поиск: ')
        self.lbl_find.move(30, 30)
        self.lbl_find.setFixedWidth(70)

        self.kandidat_find = QLineEdit(self)
        self.kandidat_find.setFocus()
        self.kandidat_find.textChanged.connect(self.kandidatFind)
        self.kandidat_find.move(100, 30)
        self.kandidat_find.setFixedWidth(230)

        self.kandidat_select = QComboBox(self)
        self.kandidat_select.move(30, 70)
        self.kandidat_select.addItems(kandidati_list)
        self.kandidat_select.setFixedWidth(300)

        # поле выбора заменяемого предмета
        self.lbl_predmet = QLabel(self)
        self.lbl_predmet.setText('Предмет: ')
        self.lbl_predmet.move(30, 110)
        self.lbl_predmet.setFixedWidth(70)

        # выбор заменяемого предмета
        pred = self.predmet
        match pred:
            case 'РУССКИЙ ЯЗЫК И ЛИТЕРАТУРА':
                self.predmet_list = ['русский язык', 'литература', 'комплексный анализ текста']
            case 'МАТЕМАТИКА':
                self.predmet_list = ['математика', 'алгебра', 'геометрия', 'практикум...']
            case 'АНГЛИЙСКИЙ ЯЗЫК':
                self.predmet_list = ['английский язык', 'китайский язык', 'немецкий язык', 'французский язык', '...']
            case 'ИНФОРМАТИКА':
                self.predmet_list = ['информатика', 'олимпиадное программирование', 'программирование', '3D-графика',
                                     'микроэлектроника', 'сетевые технологии', 'WEB - дизайн']
            case 'ФИЗИКА / АСТРОНОМИЯ':
                self.predmet_list = ['физика', 'астрономия', 'практикум...']
            case 'ХИМИЯ / ЕСТЕСТВОЗНАНИЕ':
                self.predmet_list = ['химия', 'естествознание', '', 'практикум...']
            case 'БИОЛОГИЯ':
                self.predmet_list = ['биология', 'анатомия', '', 'практикум...']
            case 'ИСТОРИЯ / ОБЩЕСТВОЗНАНИЕ':
                self.predmet_list = ['история', 'обществознание', '', 'практикум...']
            case 'ГЕОГРАФИЯ/ ОДНКНР':
                self.predmet_list = ['география', 'однкнр', '', 'практикум...']
            case 'ФИЗКУЛЬТУРА / РИТМИКА':
                self.predmet_list = ['физическая культура', 'ритмика', '', 'практикум...']
            case 'ТЕХНОЛОГИЯ':
                self.predmet_list = ['технология', 'робототехника', 'деревообработка',
                                     'авиамоделирование', 'промдизайн']
            case 'МУЗЫКА':
                self.predmet_list = ['музыка', 'практикум...']
            case 'ИЗО / ХУД.ШКОЛА':
                self.predmet_list = ['изобразительное искусство', '3D - графика', '', 'практикум...']
            case 'МУЗЕЙНАЯ ПЕДАГОГИКА':
                self.predmet_list = ['', '', '', 'практикум...']
            case 'ОБЖ':
                self.predmet_list = ['основы безопасности жизнедеятельности', '', '', 'практикум...']
            case 'ЭКОНОМИКА':
                self.predmet_list = ['экономика', '', '', 'практикум...']
            case 'КИТАЙСКИЙ ЯЗЫК / СТРАНОВЕДЕНИЕ':
                self.predmet_list = ['китайский язык', 'страноведение', '', 'практикум...']
            case 'НАЧАЛЬНАЯ ШКОЛА':
                self.predmet_list = ['математика', 'русский язык', 'литературное чтение', 'окружающий мир',
                                     'родной русский язык', 'искусство', '']

        self.predmet_select = QComboBox(self)
        self.predmet_select.move(100, 110)
        self.predmet_select.setFixedWidth(230)
        self.predmet_select.addItems(self.predmet_list)

        # поле ввода кабинета
        self.lbl_find = QLabel(self)
        self.lbl_find.setText('Кабинет: ')
        self.lbl_find.move(30, 150)
        self.lbl_find.setFixedWidth(70)

        self.kabinet = QLineEdit(self)
        self.kabinet.move(100, 150)
        self.kabinet.setFixedWidth(105)

        # поле выбора причины отсутствия
        self.lbl_prich = QLabel(self)
        self.lbl_prich.setText('Причина: ')
        self.lbl_prich.move(30, 190)
        self.lbl_prich.setFixedWidth(70)

        self.prichini = ['листок нетрудоспособности', 'отпуск без сохранения заработной платы',
                         'очередной отпуск', 'командировка']
        self.prichina_select = QComboBox(self)
        self.prichina_select.move(100, 190)
        self.prichina_select.addItems(self.prichini)
        self.prichina_select.setFixedWidth(230)

        # поле выбора ИФО
        self.lbl_ifo = QLabel(self)
        self.lbl_ifo.setText('ИФО: ')
        self.lbl_ifo.move(30, 230)
        self.lbl_ifo.setFixedWidth(70)

        self.ifo = ['СГЗ', 'ПД']
        self.ifo_select = QComboBox(self)
        self.ifo_select.move(100, 230)
        self.ifo_select.addItems(self.ifo)
        self.ifo_select.setFixedWidth(230)

        # поле выбора процента оплаты
        self.lbl_opl_proc = QLabel(self)
        self.lbl_opl_proc.setText('Оплата: ')
        self.lbl_opl_proc.move(30, 270)
        self.lbl_opl_proc.setFixedWidth(70)

        self.proc = ['100%', '50%']
        self.proc_select = QComboBox(self)
        self.proc_select.move(100, 270)
        self.proc_select.addItems(self.proc)
        self.proc_select.setFixedWidth(230)

        # кнопки отмены и добавления замены в журнал
        self.okButton = QPushButton(self)
        self.okButton.setText("Добавить")
        self.okButton.move(130, 310)
        self.okButton.setFixedWidth(100)
        self.okButton.clicked.connect(self.zamena_add)

    # вывод в выпадающий список учителей по фильтру из поля поиска
    def kandidatFind(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find.append(sotrudniki_fio[i])
        if len(sotrudniki_find) > 0:
            self.kandidat_select.clear()
            self.kandidat_select.addItems(sotrudniki_find)
        else:
            self.kandidat_select.addItems(sotrudniki_fio)

    def zamena_add(self):
        kandidat_sel = self.kandidat_select.currentText()
        kand_tab_num = kandidat_sel[:4]
        kand_fio_razd = list(map(str, kandidat_sel[6:].split()))
        kand_fio = kand_fio_razd[0] + ' ' + kand_fio_razd[1][:1] + '.' + kand_fio_razd[2][:1] + '.'
        teach_tab_num = self.selected_teacher_num_tab
        teach_fio_razd = list(map(str, self.selected_teacher_fio.split()))
        teach_fio = teach_fio_razd[0] + ' ' + teach_fio_razd[1][:1] + '.' + teach_fio_razd[2][:1] + '.'
        s_date = self.sel_date
        z_klass = self.zamena_klass.split()
        print(z_klass)
        zam_klass = z_klass[0]
        pric = self.prichina_select.currentText()
        ifo = self.ifo_select.currentText()
        oplata = self.proc_select.currentText()
        pred_sel_from_box = self.predmet_select.currentText()

        wb = load_workbook(filename=zameni_book_name)
        ws = wb['Замены']

        sheet = wb.active
        row_num = sheet.max_row

        zamena_zapis = [row_num, s_date, teach_fio, teach_tab_num, pred_sel_from_box, zam_klass, pric, kand_fio, kand_tab_num, ifo,
                        oplata]
        print(zamena_zapis)
        ws.append(zamena_zapis)
        # выравнивание столбцов B D E I F G J K по центру
        for c in 'BDEFGIJK':
            currRow = c + str(sheet.max_row)
            cell = sheet[str(currRow)]
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        wb.save(filename=zameni_book_name)
        wb.close()

        zamena_add_window.close(self)

app = QApplication(sys.argv)
app.setWindowIcon(QtGui.QIcon('../icon.png'))

window = MainWindow()
window.setGeometry(400, 200, 800, 600)
window.setWindowTitle('Составитель замен. Текущая дата: ' + today)
window.show()

app.exec()