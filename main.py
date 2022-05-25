import sys
from datetime import datetime

from PyQt6 import QtGui, QtCore, QtWidgets
from PyQt6.QtCore import QDate, Qt
from PyQt6.QtWidgets import (QApplication, QLabel, QMainWindow, QVBoxLayout, QListView, QListWidget,
                             QPushButton, QTabWidget, QWidget, QLineEdit, QComboBox, QDateEdit, QTableWidget,
                             QTableWidgetItem)

from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd

# from openpyxl.workbook import Workbook

headers = ['Фамилия', 'Имя', 'Отчество', 'Таб. номер', 'Должность', 'Дата открытия больничного листа',
           'Дата закрытия больничного листа']
bolnichniy_book_name = 'test.xlsx'
zameni_book_name = 'zameni.xlsx'

eng_chars = u"~!@#$%^&qwertyuiop[]asdfghjkl;'zxcvbnm,./QWERTYUIOP{}ASDFGHJKL:\"|ZXCVBNM<>?"
rus_chars = u"ё!\"№;%:?йцукенгшщзхъфывапролджэячсмитьбю.ЙЦУКЕНГШЩЗХЪФЫВАПРОЛДЖЭ/ЯЧСМИТЬБЮ,"
trans_table = dict(zip(eng_chars, rus_chars))

# функция исправления неправильной раскладки
def fix_input(st):
    return u''.join([trans_table.get(c, c) for c in st])

# определение текущей даты
day = datetime.now().day
if day < 10:
    day = str('0') + str(day)
month = datetime.now().month
if month < 10:
    month = str('0') + str(month)
today = str(day) + '.' + str(month) + '.' + str(datetime.now().year)
print(today)
current_day = datetime.today().weekday()

# выгрузка БД с расписанием
raspisanie = []
with open('raspisanie_done.csv', 'r') as url:
    for line in url:
        raspisanie.append(line.strip().split(';'))

# выгрузка БД с сотрудниками
sotrudniki = []
with open('sotrudniki.csv', 'r') as url:
    for line in url:
        sotrudniki.append(line.strip().split(';'))

sotrudniki_fio = []
for i in range(len(sotrudniki)):
    s = sotrudniki[i][3] + '. ' + sotrudniki[i][0] + ' ' + sotrudniki[i][1] + ' ' + sotrudniki[i][2]
    sotrudniki_fio.append(s)

class zamenaAddBtn(QPushButton):
    def __init__(self):
        QPushButton.__init__(self)
        self.setStyleSheet("background-color: rgb(255,255,255);")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Вкладки для замен")
        self.tab_index = 0
        self.tabs = QTabWidget()
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tab3 = QWidget()
        self.tab4 = QWidget()
        self.tabs.setMovable(True)
        self.tabs.addTab(self.tab1, 'Создание больничного листа')
        self.tabs.addTab(self.tab2, 'Создание замены')
        self.tabs.addTab(self.tab3, 'Журнал больничных листов')
        self.tabs.addTab(self.tab4, 'Журнал замен')
        self.setCentralWidget(self.tabs)

    ###################################################################################
    ######################### ВКЛАДКА СОЗДАНИЯ БОЛЬНИЧНОГО ЛИСТА ######################
    ###################################################################################

        # поле поиска и выбора учителя
        lbl_find = QLabel(self.tab1)
        lbl_find.setText('Поиск: ')
        lbl_find.move(30, 60)
        lbl_find.setFixedWidth(70)

        teach_find = QLineEdit(self.tab1)
        teach_find.setFocus()
        teach_find.textChanged.connect(self.teachFind)
        teach_find.move(100, 60)
        teach_find.setFixedWidth(200)

        self.teach_select = QComboBox(self.tab1)
        self.teach_select.move(30, 100)
        self.teach_select.addItems(sotrudniki_fio)
        self.teach_select.setFixedWidth(270)
        self.teach_select.currentIndex()

        # выбор даты начала замены
        self.zamena_dateStart = QDateEdit(self.tab1, calendarPopup=True)
        self.zamena_dateStart.move(100, 140)
        self.zamena_dateStart.setFixedWidth(200)
        self.zamena_dateStart.setDate(QDate.currentDate())

        lbl_zamenaStart = QLabel(self.tab1)
        lbl_zamenaStart.setText('Начало: ')
        lbl_zamenaStart.move(30, 142)
        lbl_zamenaStart.setFixedWidth(70)

        # выбор даты окончания замены
        self.zamena_dateEnd = QDateEdit(self.tab1, calendarPopup=True)
        self.zamena_dateEnd.move(100, 180)
        self.zamena_dateEnd.setFixedWidth(200)
        self.zamena_dateEnd.setDate(QDate.currentDate())

        lbl_zamenaEnd = QLabel(self.tab1)
        lbl_zamenaEnd.move(30, 182)
        lbl_zamenaEnd.setText('Окончание: ')
        lbl_zamenaEnd.setFixedWidth(70)

        # кнопки отмены и добавления замены в журнал
        okButton = QPushButton(self.tab1)
        okButton.setText("Добавить")
        okButton.move(30, 220)
        okButton.setFixedWidth(100)
        okButton.clicked.connect(self.bolnichniy_add)
        cancelButton = QPushButton(self.tab1)
        cancelButton.setText("Назад")
        cancelButton.move(200, 220)
        cancelButton.setFixedWidth(100)

    ###################################################################################
    ############################## ВКЛАДКА СОЗДАНИЯ ЗАМЕНЫ ############################
    ###################################################################################

        # выбор учителя для замены
        lbl_find_2 = QLabel(self.tab2)
        lbl_find_2.setText('Поиск: ')
        lbl_find_2.move(30, 60)
        lbl_find_2.setFixedWidth(70)

        teach_find_2 = QLineEdit(self.tab2)
        teach_find_2.setFocus()
        teach_find_2.textChanged.connect(self.teachFind_2)
        teach_find_2.move(150, 60)
        teach_find_2.setFixedWidth(200)

        self.teach_select_2 = QComboBox(self.tab2)
        self.teach_select_2.move(30, 100)
        self.teach_select_2.addItems(sotrudniki_fio)
        self.teach_select_2.setFixedWidth(270)
        self.teach_select_2.currentIndex()
        self.teach_select_2.textActivated.connect(self.zamena_teach_select)

        self.lbl_teach_select = QLabel(self.tab2)
        self.lbl_teach_select.setText('Выбранный учитель: ')
        self.lbl_teach_select.move(30, 142)
        self.lbl_teach_select.setFixedWidth(120)

        self.lbl_teach_select_2 = QLabel(self.tab2)
        self.lbl_teach_select_2.setText('Учитель еще не выбран!')
        self.lbl_teach_select_2.move(150, 142)
        self.lbl_teach_select_2.setFixedWidth(250)

        # выбор даты для замены
        lbl_zamena_2 = QLabel(self.tab2)
        lbl_zamena_2.setText('Дата замены: ')
        lbl_zamena_2.move(30, 182)
        lbl_zamena_2.setFixedWidth(70)

        self.zamena_select_2 = QDateEdit(self.tab2, calendarPopup=True)
        self.zamena_select_2.move(150, 180)
        self.zamena_select_2.setFixedWidth(200)
        self.zamena_select_2.setDate(QDate.currentDate())
        self.zamena_select_2.dateChanged.connect(self.zamena_date_select)

        self.lbl_zamena_s = QLabel(self.tab2)
        self.lbl_zamena_s.setText('Выбранная дата: ')
        self.lbl_zamena_s.move(30, 222)
        self.lbl_zamena_s.setFixedWidth(100)

        self.lbl_zamena_s_2 = QLabel(self.tab2)
        self.lbl_zamena_s_2.setText('Дата еще не выбрана!')
        self.lbl_zamena_s_2.move(150, 222)
        self.lbl_zamena_s_2.setFixedWidth(150)

        zamenaButton = QPushButton(self.tab2)
        zamenaButton.setText("Построить замены для выбранной даты")
        zamenaButton.move(30, 260)
        zamenaButton.setFixedWidth(320)
        zamenaButton.clicked.connect(lambda ch, tab = self.tab2: self.zamena_lessons_build(tab))

        self.lbl_les_1_label = QLabel(self.tab2)
        self.lbl_les_1_label.setText('1.     ' + '----------')
        self.lbl_les_1_label.move(440, 62)
        self.lbl_les_1_label.setFixedWidth(200)
        self.les_zamena_add_btn_1 = QPushButton(self.tab2)
        self.les_zamena_add_btn_1.setText("Добавить замену")
        self.les_zamena_add_btn_1.move(650, 60)
        self.les_zamena_add_btn_1.setFixedWidth(120)
        self.les_zamena_add_btn_1.setEnabled(False)
        self.les_zamena_add_btn_1.clicked.connect(lambda ch, num=1: self.zamena_add_form(num))

        self.lbl_les_2_label = QLabel(self.tab2)
        self.lbl_les_2_label.setText('2.     ' + '----------')
        self.lbl_les_2_label.move(440, 92)
        self.lbl_les_2_label.setFixedWidth(200)
        self.les_zamena_add_btn_2 = QPushButton(self.tab2)
        self.les_zamena_add_btn_2.setText("Добавить замену")
        self.les_zamena_add_btn_2.move(650, 90)
        self.les_zamena_add_btn_2.setFixedWidth(120)
        self.les_zamena_add_btn_2.setEnabled(False)
        self.les_zamena_add_btn_2.clicked.connect(lambda ch, num=2: self.zamena_add_form(num))

        self.lbl_les_3_label = QLabel(self.tab2)
        self.lbl_les_3_label.setText('3.     ' + '----------')
        self.lbl_les_3_label.move(440, 122)
        self.lbl_les_3_label.setFixedWidth(200)
        self.les_zamena_add_btn_3 = QPushButton(self.tab2)
        self.les_zamena_add_btn_3.setText("Добавить замену")
        self.les_zamena_add_btn_3.move(650, 120)
        self.les_zamena_add_btn_3.setFixedWidth(120)
        self.les_zamena_add_btn_3.setEnabled(False)
        self.les_zamena_add_btn_3.clicked.connect(lambda ch, num=3: self.zamena_add_form(num))

        self.lbl_les_4_label = QLabel(self.tab2)
        self.lbl_les_4_label.setText('4.     ' + '----------')
        self.lbl_les_4_label.move(440, 152)
        self.lbl_les_4_label.setFixedWidth(200)
        self.les_zamena_add_btn_4 = QPushButton(self.tab2)
        self.les_zamena_add_btn_4.setText("Добавить замену")
        self.les_zamena_add_btn_4.move(650, 150)
        self.les_zamena_add_btn_4.setFixedWidth(120)
        self.les_zamena_add_btn_4.setEnabled(False)
        self.les_zamena_add_btn_4.clicked.connect(lambda ch, num=4: self.zamena_add_form(num))

        self.lbl_les_5_label = QLabel(self.tab2)
        self.lbl_les_5_label.setText('5.     ' + '----------')
        self.lbl_les_5_label.move(440, 182)
        self.lbl_les_5_label.setFixedWidth(200)
        self.les_zamena_add_btn_5 = QPushButton(self.tab2)
        self.les_zamena_add_btn_5.setText("Добавить замену")
        self.les_zamena_add_btn_5.move(650, 180)
        self.les_zamena_add_btn_5.setFixedWidth(120)
        self.les_zamena_add_btn_5.setEnabled(False)
        self.les_zamena_add_btn_5.clicked.connect(lambda ch, num=5: self.zamena_add_form(num))

        self.lbl_les_6_label = QLabel(self.tab2)
        self.lbl_les_6_label.setText('6.     ' + '----------')
        self.lbl_les_6_label.move(440, 212)
        self.lbl_les_6_label.setFixedWidth(200)
        self.les_zamena_add_btn_6 = QPushButton(self.tab2)
        self.les_zamena_add_btn_6.setText("Добавить замену")
        self.les_zamena_add_btn_6.move(650, 210)
        self.les_zamena_add_btn_6.setFixedWidth(120)
        self.les_zamena_add_btn_6.setEnabled(False)
        self.les_zamena_add_btn_6.clicked.connect(lambda ch, num=6: self.zamena_add_form(num))

        self.lbl_les_7_label = QLabel(self.tab2)
        self.lbl_les_7_label.setText('7.     ' + '----------')
        self.lbl_les_7_label.move(440, 242)
        self.lbl_les_7_label.setFixedWidth(200)
        self.les_zamena_add_btn_7 = QPushButton(self.tab2)
        self.les_zamena_add_btn_7.setText("Добавить замену")
        self.les_zamena_add_btn_7.move(650, 240)
        self.les_zamena_add_btn_7.setFixedWidth(120)
        self.les_zamena_add_btn_7.setEnabled(False)
        self.les_zamena_add_btn_7.clicked.connect(lambda ch, num=7: self.zamena_add_form(num))

        self.lbl_les_8_label = QLabel(self.tab2)
        self.lbl_les_8_label.setText('8.     ' + '----------')
        self.lbl_les_8_label.move(440, 272)
        self.lbl_les_8_label.setFixedWidth(200)
        self.les_zamena_add_btn_8 = QPushButton(self.tab2)
        self.les_zamena_add_btn_8.setText("Добавить замену")
        self.les_zamena_add_btn_8.move(650, 270)
        self.les_zamena_add_btn_8.setFixedWidth(120)
        self.les_zamena_add_btn_8.setEnabled(False)
        self.les_zamena_add_btn_8.clicked.connect(lambda ch, num=8: self.zamena_add_form(num))

        self.lbl_les_9_label = QLabel(self.tab2)
        self.lbl_les_9_label.setText('9.     ' + '----------')
        self.lbl_les_9_label.move(440, 302)
        self.lbl_les_9_label.setFixedWidth(200)
        self.les_zamena_add_btn_9 = QPushButton(self.tab2)
        self.les_zamena_add_btn_9.setText("Добавить замену")
        self.les_zamena_add_btn_9.move(650, 300)
        self.les_zamena_add_btn_9.setFixedWidth(120)
        self.les_zamena_add_btn_9.setEnabled(False)
        self.les_zamena_add_btn_9.clicked.connect(lambda ch, num=9: self.zamena_add_form(num))

        self.lbl_les_10_label = QLabel(self.tab2)
        self.lbl_les_10_label.setText('10.   ' + '----------')
        self.lbl_les_10_label.move(440, 332)
        self.lbl_les_10_label.setFixedWidth(200)
        self.les_zamena_add_btn_10 = QPushButton(self.tab2)
        self.les_zamena_add_btn_10.setText("Добавить замену")
        self.les_zamena_add_btn_10.move(650, 330)
        self.les_zamena_add_btn_10.setFixedWidth(120)
        self.les_zamena_add_btn_10.setEnabled(False)
        self.les_zamena_add_btn_10.clicked.connect(lambda ch, num=10: self.zamena_add_form(num))

    ###################################################################################
    ##################### ВКЛАДКА ВЫВОДА ЖУРНАЛА БОЛЬНИЧНЫХ ЛИСТОВ ####################
    ###################################################################################

        # отслеживание нажатия на вкладку больничного листа или листа с заменами
        self.tabs.currentChanged.connect(self.currentTabNumber)

        self.bolnichniy_table = QtWidgets.QTableWidget(self.tab3)
        self.bolnichniy_table.setGeometry(30, 30, mainWindowW - 60, mainWindowH - 120)

        self.bolnichniy_table_update = QPushButton(self.tab3)
        self.bolnichniy_table_update.setText("Обновить данные")
        self.bolnichniy_table_update.move(mainWindowW//2-self.bolnichniy_table_update.width()//2, mainWindowH - 70)
        self.bolnichniy_table_update.setFixedWidth(120)
        self.bolnichniy_table_update.clicked.connect(self.bolnichniyExcelLoad)

    ###################################################################################
    ########################### ВКЛАДКА ВЫВОДА ЖУРНАЛА ЗАМЕН ##########################
    ###################################################################################

        # отслеживание нажатия на вкладку больничного листа
        self.tabs.currentChanged.connect(self.currentTabNumber)

        self.zameni_table = QtWidgets.QTableView(self.tab4)
        self.zameni_table.setGeometry(30, 30, mainWindowW - 60, mainWindowH - 120)

        self.zameni_table.horizontalHeader().resizeSection(1, 50)
        #self.zameni_table.


        self.zameni_table_update = QPushButton(self.tab4)
        self.zameni_table_update.setText("Обновить данные")
        self.zameni_table_update.move(mainWindowW//2-self.zameni_table_update.width()//2, mainWindowH - 70)
        self.zameni_table_update.setFixedWidth(120)
        self.zameni_table_update.clicked.connect(self.zameniExcelLoad)

    ###################################################################################
    ######################## ИСПОЛЬЗУЕМЫЕ ФУНКЦИИ И ПРОЦЕДУРЫ #########################
    ###################################################################################

    def bolnichniyExcelLoad(self):
        # получение данных из файла со списком больничных листов
        load_bolnichniy_file = 'test.xlsx'
        bolnichnie_data = pd.read_excel(load_bolnichniy_file, 'Учет больничных листов')
        bolnichnie_data.fillna('', inplace=True)

        self.bolnichniy_table.setRowCount(bolnichnie_data.shape[0])
        self.bolnichniy_table.setColumnCount(bolnichnie_data.shape[1])
        self.bolnichniy_table.setHorizontalHeaderLabels(bolnichnie_data.columns)

        for row in bolnichnie_data.iterrows():
            values = row[1]
            for col_index, value in enumerate(values):
                tableItem = QTableWidgetItem(str(value))
                self.bolnichniy_table.setItem(row[0], col_index, tableItem)

    def zameniExcelLoad(self):
        # получение данных из файла со списком больничных листов
        load_zameni_file = 'zameni.xlsx'
        zameni_data = pd.read_excel(load_zameni_file, 'Замены')
        zameni_data.fillna('', inplace=True)

        self.zameni_table.setRowCount(zameni_data.shape[0])
        self.zameni_table.setColumnCount(zameni_data.shape[1])
        self.zameni_table.setHorizontalHeaderLabels(zameni_data.columns)

        for row in zameni_data.iterrows():
            values = row[1]
            for col_index, value in enumerate(values):
                tableItem = QTableWidgetItem(str(value))
                self.zameni_table.setItem(row[0], col_index, tableItem)

    # определение текущей вкладки
    def currentTabNumber(self, index):
        self.tab_index = self.tabs.currentIndex()
        if self.tab_index == 2:
            self.bolnichniyExcelLoad()
        elif self.tab_index == 3:
            self.zameniExcelLoad()

    # вывод в выпадающий список учителей по фильтру из поля поиска
    def teachFind(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find.append(sotrudniki_fio[i])
        if len(sotrudniki_find) > 0:
            self.teach_select.clear()
            self.teach_select.addItems(sotrudniki_find)
        else:
            self.teach_select.addItems(sotrudniki_fio)

    def teachFind_2(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find_2 = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find_2.append(sotrudniki_fio[i])
        if len(sotrudniki_find_2) > 0:
            self.teach_select_2.clear()
            self.teach_select_2.addItems(sotrudniki_find_2)
        else:
            self.teach_select_2.addItems(sotrudniki_fio)

    def bolnichniy_add(self):
        fio_text = self.teach_select.currentText()
        id = fio_text[:4]
        for i in range(len(sotrudniki)):
            if sotrudniki[i][3] == id:
                break
        zamena_zapis = [sotrudniki[i][j] for j in range(5)]
        zamena_zapis.append(self.zamena_dateStart.text())
        zamena_zapis.append(self.zamena_dateEnd.text())
        print(zamena_zapis)

        wb = load_workbook(filename=bolnichniy_book_name)
        ws = wb['Учет больничных листов']
        ws.append(zamena_zapis)
        sheet = wb.active
        # выравнивание столбцов D E F G по центру
        for c in 'DEFG':
            currRow = c + str(sheet.max_row)
            cell = sheet[str(currRow)]
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        wb.save(filename=bolnichniy_book_name)
        wb.close()

    def zamena_date_select(self):
        selected_date = self.zamena_select_2.text()
        # print(selected_date)
        self.lbl_zamena_s_2.setText(selected_date)
        return selected_date

    def zamena_teach_select(self):
        selected_fio_text = self.teach_select_2.currentText()
        # print(selected_fio_text)
        self.lbl_teach_select_2.setText(selected_fio_text)
        return selected_fio_text[6:]

    def zamena_lessons_build(self,tab2):
        day, month, year = (int(x) for x in self.zamena_select_2.text().split('.'))
        sel_date = datetime.weekday(datetime(year, month, day))
        sel_teach = self.teach_select_2.currentText()[:4]

        if sel_date == 5 or sel_date == 6:
            self.lbl_zamena_s_2.setText('<p style="color: rgb(250, 55, 55);">ВЫБРАНА НЕВЕРНАЯ ДАТА</p>', )
        else:
            for teach_id in range(len(sotrudniki)):
                if raspisanie[teach_id][0] == sel_teach:
                    break

            for sotr_teach_id in range(len(sotrudniki)):
                if sotrudniki[sotr_teach_id][3] == sel_teach:
                    break
            print('Индекс выбранного учителя в массиве сотрудников:', teach_id)

            teach_rasp = raspisanie[teach_id][4 + sel_date * 10: 4 + (sel_date + 1) * 10]

            if teach_rasp[0] != '-':
                self.lbl_les_1_label.setText('1.     ' + teach_rasp[0])
                self.les_zamena_add_btn_1.setEnabled(True)
            else:
                self.lbl_les_1_label.setText('1.     ' + '----------')
                self.les_zamena_add_btn_1.setEnabled(False)

            if teach_rasp[1] != '-':
                self.lbl_les_2_label.setText('2.     ' + teach_rasp[1])
                self.les_zamena_add_btn_2.setEnabled(True)
            else:
                self.lbl_les_2_label.setText('2.     ' + '----------')
                self.les_zamena_add_btn_2.setEnabled(False)

            if teach_rasp[2] != '-':
                self.lbl_les_3_label.setText('3.     ' + teach_rasp[2])
                self.les_zamena_add_btn_3.setEnabled(True)
            else:
                self.lbl_les_3_label.setText('3.     ' + '----------')
                self.les_zamena_add_btn_3.setEnabled(False)

            if teach_rasp[3] != '-':
                self.lbl_les_4_label.setText('4.     ' + teach_rasp[3])
                self.les_zamena_add_btn_4.setEnabled(True)
            else:
                self.lbl_les_4_label.setText('4.     ' + '----------')
                self.les_zamena_add_btn_4.setEnabled(False)

            if teach_rasp[4] != '-':
                self.lbl_les_5_label.setText('5.     ' + teach_rasp[4])
                self.les_zamena_add_btn_5.setEnabled(True)
            else:
                self.lbl_les_5_label.setText('5.     ' + '----------')
                self.les_zamena_add_btn_5.setEnabled(False)

            if teach_rasp[5] != '-':
                self.lbl_les_6_label.setText('6.     ' + teach_rasp[5])
                self.les_zamena_add_btn_6.setEnabled(True)
            else:
                self.lbl_les_6_label.setText('6.     ' + '----------')
                self.les_zamena_add_btn_6.setEnabled(False)

            if teach_rasp[6] != '-':
                self.lbl_les_7_label.setText('7.     ' + teach_rasp[6])
                self.les_zamena_add_btn_7.setEnabled(True)
            else:
                self.lbl_les_7_label.setText('7.     ' + '----------')
                self.les_zamena_add_btn_7.setEnabled(False)

            if teach_rasp[7] != '-':
                self.lbl_les_8_label.setText('8.     ' + teach_rasp[7])
                self.les_zamena_add_btn_8.setEnabled(True)
            else:
                self.lbl_les_8_label.setText('8.     ' + '----------')
                self.les_zamena_add_btn_8.setEnabled(False)

            if teach_rasp[8] != '-':
                self.lbl_les_9_label.setText('9.     ' + teach_rasp[8])
                self.les_zamena_add_btn_9.setEnabled(True)
            else:
                self.lbl_les_9_label.setText('9.     ' + '----------')
                self.les_zamena_add_btn_9.setEnabled(False)

            if teach_rasp[9] != '-':
                self.lbl_les_10_label.setText('10.   ' + teach_rasp[9])
                self.les_zamena_add_btn_10.setEnabled(True)
            else:
                self.lbl_les_10_label.setText('10.   ' + '----------')
                self.les_zamena_add_btn_10.setEnabled(False)

    def zamena_add_form(self, i):
        num_les = i
        day, month, year = (int(x) for x in self.zamena_select_2.text().split('.'))
        date = datetime(year, month, day)
        sel_date = datetime.weekday(datetime(year, month, day))
        sel_teach_all = self.teach_select_2.currentText()
        sel_teach = self.teach_select_2.currentText()[:4]

        if sel_date == 5 or sel_date == 6:
            self.lbl_zamena_s_2.setText('<p style="color: rgb(250, 55, 55);">ВЫБРАНА НЕВЕРНАЯ ДАТА</p>', )
        else:
            for teach_id in range(len(sotrudniki)):
                if raspisanie[teach_id][0] == sel_teach:
                    break

            for sotr_teach_id in range(len(sotrudniki)):
                if sotrudniki[sotr_teach_id][3] == sel_teach:
                    break
            print('Индекс выбранного учителя в массиве сотрудников:', teach_id)

            # вывод заменяемого учителя и расписание его уроков на этот день
            sel_teach_num_les = []
            for i in range(4 + sel_date * 10, 4 + (sel_date + 1) * 10):
                if raspisanie[teach_id][i] != '-':
                    # сохранение в массив sel_teach_num_les уроков для замены (окна пропущены)
                    sel_teach_num_les.append(i)
            print('Номера уроков для замены:', *sel_teach_num_les)
            self.sel_predmet = raspisanie[teach_id][2]
            # вывод учителей ТОГО ЖЕ предмета
            kandidati = []
            self.send_teach_id = teach_id

            for j in range(len(raspisanie)):
                for i in sel_teach_num_les:
                    if raspisanie[j][2] == self.sel_predmet and raspisanie[j][0] != sel_teach \
                            and raspisanie[j][i] == '-':
                        kand_temp = str(i - 3 - sel_date * 10) + ';' + str(raspisanie[j][0]) + ';' \
                                    + str(raspisanie[j][1]) + ';' + str(raspisanie[teach_id][i])
                        kandidati.append(kand_temp)

            # вывод учителей ОСТАЛЬНЫХ предметов
            for j in range(len(raspisanie)):
                for i in sel_teach_num_les:
                    if raspisanie[j][2] != self.sel_predmet and raspisanie[j][i] == '-':
                        kand_temp = str(i - 3 - sel_date * 10) + ';' + str(raspisanie[j][0]) + ';' + \
                                    str(raspisanie[j][1]) + ';' + str(raspisanie[teach_id][i])
                        kandidati.append(kand_temp)

            kandidati = sorted(kandidati, key=lambda row: row[0])
            n_lessons = 10 - raspisanie[teach_id][4 + sel_date * 10: 4 + (sel_date + 1) * 10].count('-')
            print('Количество заменяемых уроков:', n_lessons)
            print('Количество кандидатов для замены ВСЕХ уроков:', len(kandidati))

            self.win_zamena = zamena_add_window(teach_id, num_les, sel_teach_all, kandidati,
                                                date, sel_date, self.sel_predmet)
            self.win_zamena.show()

class zamena_add_window(QWidget):
    def __init__(self, teach_id, num_les, sel_teach, kandidati, date, sel_weekday, sel_predmet):
        super().__init__()
        self.selected_teacher_fio = sel_teach[6:]
        self.selected_teacher_num_tab = sel_teach[:4]
        self.setWindowTitle('Окно добавления замены')
        self.setFixedSize(360, 340)
        year, month, day = str(date)[:10].split('-')
        self.sel_date = day + '.' + month + '.' + year
        self.zamena_less = 4 + int(sel_weekday) * 10 + int(num_les)-1
        self.zamena_klass = raspisanie[teach_id][self.zamena_less]
        self.predmet = sel_predmet
        print('Выбранная дата:', self.sel_date)
        print('Выбранный учитель ТАБ №:', self.selected_teacher_num_tab)
        print('Выбранный учитель ФИО:', self.selected_teacher_fio)
        print('Кафедра на котором работает учитель:', self.predmet)
        print('Номер дня недели:', sel_weekday + 1)
        print('Номер заменяемого урока:', num_les)
        print('Индекс столбца заменяемого урока:', self.zamena_less)
        print('Класс для которого выставляется замена на данном уроке:', self.zamena_klass)

        kandidati_list = []
        for i in range(len(kandidati)):
            sel_les, num_tab_kand, fio_kand, zamena_les = list(map(str, kandidati[i].split(';')))
            kandidati_list.append(num_tab_kand + '. ' + fio_kand)

        # поле поиска и выбора учителя
        self.lbl_find = QLabel(self)
        self.lbl_find.setText('Поиск: ')
        self.lbl_find.move(30, 30)
        self.lbl_find.setFixedWidth(70)

        self.kandidat_find = QLineEdit(self)
        self.kandidat_find.setFocus()
        self.kandidat_find.textChanged.connect(self.kandidatFind)
        self.kandidat_find.move(100, 30)
        self.kandidat_find.setFixedWidth(230)

        self.kandidat_select = QComboBox(self)
        self.kandidat_select.move(30, 70)
        self.kandidat_select.addItems(kandidati_list)
        self.kandidat_select.setFixedWidth(300)

        # поле выбора заменяемого предмета
        self.lbl_predmet = QLabel(self)
        self.lbl_predmet.setText('Предмет: ')
        self.lbl_predmet.move(30, 110)
        self.lbl_predmet.setFixedWidth(70)

        # выбор заменяемого предмета
        pred = self.predmet
        match pred:
            case 'РУССКИЙ ЯЗЫК И ЛИТЕРАТУРА':
                self.predmet_list = ['русский язык', 'литература', 'комплексный анализ текста']
            case 'МАТЕМАТИКА':
                self.predmet_list = ['математика', 'алгебра', 'геометрия', 'практикум...']
            case 'АНГЛИЙСКИЙ ЯЗЫК':
                self.predmet_list = ['английский язык', 'китайский язык', 'немецкий язык', 'французский язык', '...']
            case 'ИНФОРМАТИКА':
                self.predmet_list = ['информатика', 'олимпиадное программирование', 'программирование', '3D-графика',
                                     'микроэлектроника', 'сетевые технологии', 'WEB - дизайн']
            case 'ФИЗИКА / АСТРОНОМИЯ':
                self.predmet_list = ['физика', 'астрономия', 'практикум...']
            case 'ХИМИЯ / ЕСТЕСТВОЗНАНИЕ':
                self.predmet_list = ['химия', 'естествознание', '', 'практикум...']
            case 'БИОЛОГИЯ':
                self.predmet_list = ['биология', 'анатомия', '', 'практикум...']
            case 'ИСТОРИЯ / ОБЩЕСТВОЗНАНИЕ':
                self.predmet_list = ['история', 'обществознание', '', 'практикум...']
            case 'ГЕОГРАФИЯ/ ОДНКНР':
                self.predmet_list = ['география', 'однкнр', '', 'практикум...']
            case 'ФИЗКУЛЬТУРА / РИТМИКА':
                self.predmet_list = ['физическая культура', 'ритмика', '', 'практикум...']
            case 'ТЕХНОЛОГИЯ':
                self.predmet_list = ['технология', 'робототехника', 'деревообработка',
                                     'авиамоделирование', 'промдизайн']
            case 'МУЗЫКА':
                self.predmet_list = ['музыка', 'практикум...']
            case 'ИЗО / ХУД.ШКОЛА':
                self.predmet_list = ['изобразительное искусство', '3D - графика', '', 'практикум...']
            case 'МУЗЕЙНАЯ ПЕДАГОГИКА':
                self.predmet_list = ['', '', '', 'практикум...']
            case 'ОБЖ':
                self.predmet_list = ['основы безопасности жизнедеятельности', '', '', 'практикум...']
            case 'ЭКОНОМИКА':
                self.predmet_list = ['экономика', '', '', 'практикум...']
            case 'КИТАЙСКИЙ ЯЗЫК / СТРАНОВЕДЕНИЕ':
                self.predmet_list = ['китайский язык', 'страноведение', '', 'практикум...']
            case 'НАЧАЛЬНАЯ ШКОЛА':
                self.predmet_list = ['математика', 'русский язык', 'литературное чтение', 'окружающий мир',
                                     'родной русский язык', 'искусство', '']

        self.predmet_select = QComboBox(self)
        self.predmet_select.move(100, 110)
        self.predmet_select.setFixedWidth(230)
        self.predmet_select.addItems(self.predmet_list)

        # поле ввода кабинета
        self.lbl_find = QLabel(self)
        self.lbl_find.setText('Кабинет: ')
        self.lbl_find.move(30, 150)
        self.lbl_find.setFixedWidth(70)

        self.kabinet = QLineEdit(self)
        self.kabinet.move(100, 150)
        self.kabinet.setFixedWidth(105)

        # поле выбора причины отсутствия
        self.lbl_prich = QLabel(self)
        self.lbl_prich.setText('Причина: ')
        self.lbl_prich.move(30, 190)
        self.lbl_prich.setFixedWidth(70)

        self.prichini = ['листок нетрудоспособности', 'отпуск без сохранения заработной платы',
                         'очередной отпуск', 'командировка']
        self.prichina_select = QComboBox(self)
        self.prichina_select.move(100, 190)
        self.prichina_select.addItems(self.prichini)
        self.prichina_select.setFixedWidth(230)

        # поле выбора ИФО
        self.lbl_ifo = QLabel(self)
        self.lbl_ifo.setText('ИФО: ')
        self.lbl_ifo.move(30, 230)
        self.lbl_ifo.setFixedWidth(70)

        self.ifo = ['СГЗ', 'ПД']
        self.ifo_select = QComboBox(self)
        self.ifo_select.move(100, 230)
        self.ifo_select.addItems(self.ifo)
        self.ifo_select.setFixedWidth(230)

        # поле выбора процента оплаты
        self.lbl_opl_proc = QLabel(self)
        self.lbl_opl_proc.setText('Оплата: ')
        self.lbl_opl_proc.move(30, 270)
        self.lbl_opl_proc.setFixedWidth(70)

        self.proc = ['100%', '50%']
        self.proc_select = QComboBox(self)
        self.proc_select.move(100, 270)
        self.proc_select.addItems(self.proc)
        self.proc_select.setFixedWidth(230)

        # кнопки отмены и добавления замены в журнал
        self.okButton = QPushButton(self)
        self.okButton.setText("Добавить")
        self.okButton.move(130, 310)
        self.okButton.setFixedWidth(100)
        self.okButton.clicked.connect(self.zamena_add)

    # вывод в выпадающий список учителей по фильтру из поля поиска
    def kandidatFind(self, text):
        fioSrez = fix_input(text.capitalize())
        lenFioSrez = len(fioSrez)
        sotrudniki_find = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][6:6 + lenFioSrez] == fioSrez:
                sotrudniki_find.append(sotrudniki_fio[i])
        if len(sotrudniki_find) > 0:
            self.kandidat_select.clear()
            self.kandidat_select.addItems(sotrudniki_find)
        else:
            self.kandidat_select.addItems(sotrudniki_fio)

    def zamena_add(self):
        kandidat_sel = self.kandidat_select.currentText()
        kand_tab_num = kandidat_sel[:4]
        kand_fio_razd = list(map(str, kandidat_sel[6:].split()))
        kand_fio = kand_fio_razd[0] + ' ' + kand_fio_razd[1][:1] + '.' + kand_fio_razd[2][:1] + '.'
        teach_tab_num = str(self.selected_teacher_num_tab)
        teach_fio_razd = list(map(str, self.selected_teacher_fio.split()))
        teach_fio = teach_fio_razd[0] + ' ' + teach_fio_razd[1][:1] + '.' + teach_fio_razd[2][:1] + '.'
        s_date = self.sel_date
        z_klass = self.zamena_klass.split()
        print(z_klass)
        zam_klass = z_klass[0]
        pric = self.prichina_select.currentText()
        ifo = self.ifo_select.currentText()
        oplata = self.proc_select.currentText()
        pred_sel_from_box = self.predmet_select.currentText()

        wb = load_workbook(filename=zameni_book_name)
        ws = wb['Замены']

        sheet = wb.active
        row_num = sheet.max_row

        zamena_zapis = [row_num, s_date, teach_fio, teach_tab_num, pred_sel_from_box, zam_klass, pric, kand_fio, kand_tab_num, ifo,
                        oplata]
        print(zamena_zapis)
        ws.append(zamena_zapis)

        # выравнивание столбцов B D E I F G J K по центру
        for c in 'BDEFGIJK':
            currRow = c + str(sheet.max_row)
            cell = sheet[str(currRow)]
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        wb.save(filename=zameni_book_name)
        wb.close()

        zamena_add_window.close(self)

app = QApplication(sys.argv)
app.setWindowIcon(QtGui.QIcon('icon.png'))

mainWindowW = 1300
mainWindowH = 600

window = MainWindow()
windowW = window.frameSize().width()
windowH = window.frameSize().height()
print(windowW, windowH)
window.setGeometry((int(1920/1.25)-mainWindowW)//2, (int(1200/1.25)-mainWindowH)//2, mainWindowW, mainWindowH)
window.setWindowTitle('Составитель замен. Текущая дата: ' + today)
window.show()

app.exec()