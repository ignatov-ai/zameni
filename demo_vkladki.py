import sys
import pandas as pd
import os
from openpyxl import load_workbook

from PyQt6.QtWidgets import (
    QApplication,
    QLabel,
    QMainWindow,
    QPushButton,
    QTabWidget,
    QWidget, QLineEdit, QComboBox, QDateEdit, QHBoxLayout, QVBoxLayout,
)

from PyQt6.QtCore import QDate


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


sotrudniki = []
with open('sotrudniki.csv','r') as url:
  for line in url:
    sotrudniki.append(line.strip().split(';'))

sotrudniki_fio = []
for i in range(len(sotrudniki)):
    s = sotrudniki[i][0]+' '+sotrudniki[i][1]+' '+sotrudniki[i][2]
    sotrudniki_fio.append(s)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Вкладки для замен")

        self.tabs = QTabWidget()
        tab1 = QWidget()
        tab2 = QWidget()
        tab3 = QWidget()
        self.tabs.setMovable(True)
        self.tabs.addTab(tab1, 'Создание больничного листа')
        self.tabs.addTab(tab2, '2')
        self.tabs.addTab(tab3, '3')
        self.setCentralWidget(self.tabs)

        # поле поиска и выбора учителя
        lbl_find = QLabel(tab1)
        lbl_find.setText('Поиск: ')
        lbl_find.move(30,60)
        lbl_find.setFixedWidth(70)

        teach_find = QLineEdit(tab1)
        teach_find.textChanged.connect(self.teachFind)
        teach_find.move(100,60)
        teach_find.setFixedWidth(200)

        self.teach_select = QComboBox(tab1)
        self.teach_select.move(30,100)
        self.teach_select.addItems(sotrudniki_fio)
        self.teach_select.setFixedWidth(270)
        self.teach_select.currentIndex()

        # выбор даты начала замены
        zamena_dateStart = QDateEdit(tab1, calendarPopup=True)
        zamena_dateStart.move(100,140)
        zamena_dateStart.setFixedWidth(200)
        zamena_dateStart.setDate(QDate.currentDate())

        lbl_zamenaStart = QLabel(tab1)
        lbl_zamenaStart.setText('Начало: ')
        lbl_zamenaStart.move(30,142)
        lbl_zamenaStart.setFixedWidth(70)

        # выбор даты окончания замены
        zamena_dateEnd = QDateEdit(tab1, calendarPopup=True)
        zamena_dateEnd.move(100,180)
        zamena_dateEnd.setFixedWidth(200)
        zamena_dateEnd.setDate(QDate.currentDate())

        lbl_zamenaEnd = QLabel(tab1)
        lbl_zamenaEnd.move(30,182)
        lbl_zamenaEnd.setText('Окончание: ')
        lbl_zamenaEnd.setFixedWidth(70)

        # кнопки отмены и добавления замены в журнал
        okButton = QPushButton(tab1)
        okButton.setText("Добавить")
        okButton.move(30,220)
        okButton.setFixedWidth(100)
        okButton.clicked.connect(self.zamena_add)
        #okButton.addAction(self.zamena_add)
        cancelButton = QPushButton(tab1)
        cancelButton.setText("Назад")
        cancelButton.move(200,220)
        cancelButton.setFixedWidth(100)

        hBtn = QHBoxLayout()
        hBtn.addWidget(okButton)
        hBtn.addWidget(cancelButton)

        # вывод в выпадающий список учителей по фильтру из поля поиска
    def teachFind(self, text):
        fioSrez = text.capitalize()
        lenFioSrez = len(fioSrez)
        sotrudniki_find = []
        for i in range(len(sotrudniki_fio)):
            if sotrudniki_fio[i][:lenFioSrez] == fioSrez:
                sotrudniki_find.append(sotrudniki_fio[i])
        if len(sotrudniki_find) > 0:
            self.teach_select.clear()
            self.teach_select.addItems(sotrudniki_find)
        else:
            self.teach_select.addItems(sotrudniki_fio)

    def zamena_add(self):
        id = self.teach_select.currentIndex()
        print(sotrudniki[id])



        #zameni_book = xlsxwriter.workbook('zameni_book.xlsx')
        #zameni_book = zameni_book.add_worksheet()
        new_zapis = [[sotrudniki[id][0], sotrudniki[id][1], sotrudniki[id][2], sotrudniki[id][3], sotrudniki[id][4]]]
        df = pd.DataFrame(new_zapis)
        writer = pd.ExcelWriter('test.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace')
        #df.to_excel(writer, sheet_name='zameni', startrow=writer.sheets['zameni'].max_row, index=False, header=False)
        #writer.save()

        append_df_to_excel('test.xlsx', df, sheet_name='zameni', startrow=len(df)+1, header=None, index=False)


app = QApplication(sys.argv)

window = MainWindow()
window.setGeometry(400, 200, 800, 600)
window.show()

app.exec()